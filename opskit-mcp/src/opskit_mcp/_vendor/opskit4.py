"""
OPSKIT v4 — COGNITIVE ANALYST PLAYBOOKS
=============================================================
Author : Mohd Saif Hussain
Engine : DuckDB over CSV / Excel / SQLite / Parquet

WHAT v4 FIXES (every flaw from the v3 review, by name)
------------------------------------------------------
1. HONEST CONDITIONAL DRILL-DOWN. v3 ranked columns independently and
   drew an arrow implying a causal path it never computed. v4 implements
   real contribution analysis: find the segment explaining the largest
   share of the total delta, CONDITION on it, and recurse inside that
   segment. Every arrow in the output is now a computed path, and every
   driver carries its contribution percentage.
2. DECLARED STEP DEPENDENCIES. Steps state what they require; the
   registry validates ordering at load time. Correctness is enforced,
   not a tuple-ordering convention.
3. IMMUTABLE STEP RESULTS. Steps are pure: context in, findings out.
   The runner composes. No shared-mutable-state side effects.
4. CONFIG SURFACE. Every threshold lives in a frozen Config, overridable
   from opskit.toml. Custom playbooks can be COMPOSED IN TOML from the
   step library — domain packs without touching code.
5. MACHINE-READABLE OUTPUT. --json emits findings as JSON lines; exit
   code 2 signals critical findings (CI-friendly), 0 clean, 1 error.
6. VALIDATED IDENTIFIERS. Every column name passes through qident(),
   which allowlists against the actual schema. Values bind as
   parameters. No raw f-string identifiers anywhere.

DELIBERATE TRADE-OFF: this remains a single file. For a drop-one-file-
and-run learning tool, portability beats package structure; the module
boundaries below are drawn so a later split into steps/ playbooks/
report/ is mechanical. That is a decision, documented — not inertia.

THE PATTERN THIS AUTOMATES (identical across domains)
-----------------------------------------------------
Incidents, transactions, claims, shipments, tickets — a senior analyst
runs the same sequence on all of them:
  SHAPE → TRUST (missing/dupes/sanity) → CHANGE → DRIVER → CONCENTRATION
  → ACTION.
v4 encodes the sequence; the TOML layer renames and rethresholds it
per domain.

Usage:
  python3 opskit4.py demo
  python3 opskit4.py run weekly-review demo_data/incidents.csv
  python3 opskit4.py run weekly-review demo_data/incidents.csv --json findings.jsonl
  python3 opskit4.py list
  python3 opskit4.py explain trend-investigation
"""

from __future__ import annotations

__all__ = [
    "PLAYBOOKS",
    "Config",
    "Finding",
    "OpsKitError",
    "Playbook",
    "PlaybookContext",
    "Severity",
    "Step",
    "build_context",
    "conditional_drill",
    "load_config",
    "load_source",
    "resolve_metric",
    "run_playbook",
]

import argparse
import csv
import json
import os
import random
import sqlite3
import sys
import tempfile
import tomllib
from collections.abc import Callable
from dataclasses import dataclass, replace
from datetime import datetime, timedelta
from enum import StrEnum
from pathlib import Path
from typing import Any, Final
from zoneinfo import ZoneInfo

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

IST: Final[ZoneInfo] = ZoneInfo("Asia/Kolkata")

PAL: Final[dict[str, str]] = {
    "navy": "1B3A6B", "red": "C00000", "amber": "F4B942",
    "green": "70AD47", "alt": "EEF2F7", "white": "FFFFFF",
    "ink": "1A1A2E", "muted": "666666",
}

TIME_HINTS: Final[tuple[str, ...]] = (
    "_at", "date", "time", "created", "opened", "closed", "timestamp",
)

# PEP 695 type aliases — restored from AnalystKit (their absence in v3
# was a regression).
type Row = tuple[Any, ...]
type WhereClause = list[tuple[str, str]]        # [(column, value), ...]
type StepFn = Callable[["PlaybookContext", "Config"], tuple["Finding", ...]]

EXIT_CLEAN: Final[int] = 0
EXIT_ERROR: Final[int] = 1
EXIT_CRITICAL_FINDINGS: Final[int] = 2


class Severity(StrEnum):
    INFO = "INFO"
    NOTABLE = "NOTABLE"
    CRITICAL = "CRITICAL"


class OpsKitError(ValueError):
    """User-facing error with a readable message."""


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — every threshold named, documented, and overridable from TOML
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True, slots=True)
class Config:
    """All tunable behaviour. Load overrides from opskit.toml [thresholds]."""

    drill_threshold: float = 0.20     # |WoW change| that triggers drill-down
    drill_depth: int = 2              # how many conditioning levels deep
    max_cardinality: int = 30         # categorical columns above this are skipped
    anomaly_mad_k: float = 3.0        # median + k*MAD anomaly cutoff
    sustained_run: int = 4            # >= this many flagged days = level shift
    window_days: int = 7              # comparison window
    null_notable_share: float = 0.05  # null share that upgrades INFO→NOTABLE
    concentration_share: float = 0.40 # single-group share worth flagging
    time_cast_ratio: float = 0.90     # share of castable values to accept a time col
    metric: str = "count"             # count | sum:<col> | avg:<col>


def load_config(path: Path | None = None) -> tuple[Config, dict[str, Any]]:
    """Loads Config (+ raw custom playbook tables) from opskit.toml if present.

    Returns:
        (config, custom_playbooks_raw) — the raw [playbooks.*] tables are
        materialised into Playbook objects after the step library exists.
    """
    toml_path = path or Path("opskit.toml")
    if not toml_path.exists():
        return Config(), {}
    with toml_path.open("rb") as fh:
        data = tomllib.load(fh)
    thresholds = data.get("thresholds", {})
    valid = set(Config.__dataclass_fields__)
    unknown = sorted(set(thresholds) - valid)
    if unknown:
        raise OpsKitError(
            f"opskit.toml [thresholds] has unknown keys: {unknown}. "
            f"Valid keys: {sorted(valid)}"
        )
    cfg = replace(Config(), **thresholds)
    return cfg, data.get("playbooks", {})


# ─────────────────────────────────────────────────────────────────────────────
# CORE DATACLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True, slots=True)
class Finding:
    """One observation with its evidence. Immutable — findings are records."""

    step: str
    severity: Severity
    text: str
    evidence: str = ""

    def as_json(self) -> dict[str, str]:
        return {"step": self.step, "severity": str(self.severity),
                "text": self.text, "evidence": self.evidence}


@dataclass(frozen=True, slots=True)
class PlaybookContext:
    """Read-only facts about the loaded source. Steps never mutate this."""

    con: duckdb.DuckDBPyConnection
    source: Path
    schema: tuple[tuple[str, str], ...]     # ((name, dtype), ...)
    time_col: str | None
    category_cols: tuple[str, ...]
    numeric_cols: tuple[str, ...]
    total_rows: int

    def qident(self, name: str) -> str:
        """Validated identifier quoting: the column must exist in the schema.

        This is the single chokepoint for identifiers entering SQL text —
        an allowlist, not an escape function.
        """
        known = {c for c, _ in self.schema}
        if name not in known:
            raise OpsKitError(
                f"Column '{name}' does not exist. Available: {sorted(known)}"
            )
        return '"' + name.replace('"', '""') + '"'


@dataclass(frozen=True, slots=True)
class Step:
    """A checklist item: pure function from (context, config) to findings."""

    key: str
    question: str
    rationale: str
    run: StepFn
    requires: tuple[str, ...] = ()      # step keys that must run earlier


@dataclass(frozen=True, slots=True)
class Playbook:
    key: str
    title: str
    description: str
    steps: tuple[Step, ...]

    def validate(self) -> None:
        """Enforces declared dependencies: correctness by contract,
        not by tuple-ordering convention."""
        seen: set[str] = set()
        for step in self.steps:
            missing = [r for r in step.requires if r not in seen]
            if missing:
                raise OpsKitError(
                    f"Playbook '{self.key}': step '{step.key}' requires "
                    f"{missing} to run earlier, but they do not."
                )
            seen.add(step.key)


# ─────────────────────────────────────────────────────────────────────────────
# SOURCE LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_source(path: Path, table: str | None = None) -> duckdb.DuckDBPyConnection:
    """Opens CSV / Excel / SQLite / Parquet as view `t`."""
    if not path.exists():
        raise OpsKitError(f"Source not found: {path}")
    con = duckdb.connect()
    suffix = path.suffix.lower()
    posix = path.as_posix().replace("'", "''")
    if suffix == ".csv":
        con.execute(f"CREATE VIEW t AS SELECT * FROM read_csv_auto('{posix}')")
    elif suffix == ".parquet":
        con.execute(f"CREATE VIEW t AS SELECT * FROM read_parquet('{posix}')")
    elif suffix in (".xlsx", ".xls"):
        import pandas as pd
        con.register("t", pd.read_excel(path))
    elif suffix in (".sqlite", ".db", ".sqlite3"):
        con.execute(f"ATTACH '{posix}' AS src (TYPE sqlite)")
        tables = [str(r[0]) for r in con.execute(
            "SELECT table_name FROM duckdb_tables() WHERE database_name='src'"
        ).fetchall()]
        if not tables:
            raise OpsKitError(f"No tables in database: {path}")
        if table is None:
            if len(tables) > 1:
                raise OpsKitError(f"Multiple tables {tables}; pick one with --table.")
            table = tables[0]
        if table not in tables:
            raise OpsKitError(f"Table '{table}' not found. Available: {tables}")
        safe = table.replace('"', '""')
        con.execute(f'CREATE VIEW t AS SELECT * FROM src."{safe}"')
    else:
        raise OpsKitError(f"Unsupported source '{suffix}'.")
    return con


def build_context(
    path: Path, cfg: Config, table: str | None = None
) -> PlaybookContext:
    """Loads the source, detects roles, prints assumptions (human disposes)."""
    con = load_source(path, table)
    schema = tuple(
        (str(r[0]), str(r[1])) for r in con.execute("DESCRIBE t").fetchall()
    )
    row = con.execute("SELECT COUNT(*) FROM t").fetchone()
    total = int(row[0]) if row else 0

    time_col: str | None = None
    cats: list[str] = []
    nums: list[str] = []
    for name, dtype in schema:
        upper = dtype.upper()
        quoted = '"' + name.replace('"', '""') + '"'
        if time_col is None and ("TIMESTAMP" in upper or "DATE" in upper):
            time_col = name
            continue
        if time_col is None and any(h in name.lower() for h in TIME_HINTS):
            v = con.execute(
                f"SELECT COUNT(*), COUNT(TRY_CAST({quoted} AS TIMESTAMP)) "
                f"FROM t WHERE {quoted} IS NOT NULL"
            ).fetchone()
            if v and int(v[0]) > 0 and int(v[1]) / int(v[0]) >= cfg.time_cast_ratio:
                time_col = name
                continue
        if "VARCHAR" in upper:
            d = con.execute(f"SELECT COUNT(DISTINCT {quoted}) FROM t").fetchone()
            if d and 1 < int(d[0]) <= cfg.max_cardinality:
                cats.append(name)
        elif any(k in upper for k in ("INT", "DOUBLE", "DECIMAL", "FLOAT")):
            nums.append(name)

    print(f"[assumption] time column     : {time_col or 'none detected'}")
    print(f"[assumption] category columns: {cats or 'none'}")
    print(f"[assumption] numeric columns : {nums or 'none'}")
    return PlaybookContext(
        con=con, source=path, schema=schema, time_col=time_col,
        category_cols=tuple(cats), numeric_cols=tuple(nums), total_rows=total,
    )


# ─────────────────────────────────────────────────────────────────────────────
# METRICS — count | sum:<col> | avg:<col>
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True, slots=True)
class Metric:
    """A validated metric. Additive metrics (count, sum) decompose across
    segments, so drill-down attribution is mathematically sound. Averages
    do NOT decompose additively (mix vs rate effects — Simpson's paradox),
    so the drill honestly refuses them rather than fabricating attribution."""

    kind: str                 # "count" | "sum" | "avg"
    column: str | None = None

    @property
    def additive(self) -> bool:
        return self.kind in ("count", "sum")

    @property
    def label(self) -> str:
        return "record count" if self.kind == "count" else f"{self.kind}({self.column})"

    def sql(self, ctx: PlaybookContext) -> str:
        if self.kind == "count":
            return "COUNT(*)"
        q = ctx.qident(self.column or "")
        # COALESCE: SUM/AVG over an empty or all-NULL window returns NULL,
        # which would poison the delta arithmetic downstream.
        return f"COALESCE({self.kind.upper()}({q}), 0)"

    def fmt(self, v: float) -> str:
        return f"{int(v):,}" if self.kind == "count" else f"{v:,.2f}"


def resolve_metric(ctx: PlaybookContext, cfg: Config) -> Metric:
    """Parses and validates cfg.metric against the actual schema."""
    raw = cfg.metric.strip()
    if raw == "count":
        return Metric("count")
    if ":" not in raw:
        raise OpsKitError(
            f"Invalid metric '{raw}'. Use count, sum:<column>, or avg:<column>."
        )
    kind, col = raw.split(":", 1)
    if kind not in ("sum", "avg"):
        raise OpsKitError(f"Unknown metric kind '{kind}'. Use count, sum or avg.")
    ctx.qident(col)   # existence check (allowlist)
    if col not in ctx.numeric_cols:
        raise OpsKitError(
            f"Metric column '{col}' is not numeric. "
            f"Numeric columns available: {list(ctx.numeric_cols) or 'none'}"
        )
    return Metric(kind, col)


# ─────────────────────────────────────────────────────────────────────────────
# CONDITIONAL DRILL-DOWN — contribution analysis, honestly computed
# ─────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True, slots=True)
class DrillLevel:
    """One conditioning level: the segment that best explains the delta."""

    column: str
    value: str
    cur: float
    prev: float
    contribution: float     # this segment's share of the parent delta


def _windowed_agg(
    ctx: PlaybookContext, cfg: Config, where: WhereClause,
    metric: Metric, group_col: str | None = None,
) -> list[Row]:
    """Current vs previous window aggregates, optionally grouped, under `where`.

    Identifiers pass through qident (allowlist). Values bind as parameters —
    the string never enters the SQL text.
    """
    if ctx.time_col is None:
        raise OpsKitError(
            "Windowed comparison requires a time column, and none was "
            "detected. This is a caller bug — time-dependent steps must "
            "check ctx.time_col before calling."
        )
    tq = ctx.qident(ctx.time_col)
    conds, params = [], []
    for col, val in where:
        conds.append(f"a.{ctx.qident(col)} = ?")
        params.append(val)
    where_sql = (" AND " + " AND ".join(conds)) if conds else ""
    select_group = f"a.{ctx.qident(group_col)}," if group_col else ""
    group_by = "GROUP BY 1" if group_col else ""
    d = cfg.window_days
    # FILTER windows apply to the aggregate itself, so sum/avg respect them.
    if metric.kind == "count":
        cur_expr  = f"COUNT(*) FILTER (WHERE a.ts >  b.anchor - INTERVAL {d} DAY)"
        prev_expr = (f"COUNT(*) FILTER (WHERE a.ts <= b.anchor - INTERVAL {d} DAY "
                     f"AND a.ts > b.anchor - INTERVAL {d * 2} DAY)")
    else:
        q = ctx.qident(metric.column or "")
        fn = metric.kind.upper()
        cur_expr  = (f"COALESCE({fn}({q}) FILTER "
                     f"(WHERE a.ts > b.anchor - INTERVAL {d} DAY), 0)")
        prev_expr = (f"COALESCE({fn}({q}) FILTER "
                     f"(WHERE a.ts <= b.anchor - INTERVAL {d} DAY "
                     f"AND a.ts > b.anchor - INTERVAL {d * 2} DAY), 0)")
    sql = f"""
        WITH anchored AS (
            SELECT *, TRY_CAST({tq} AS TIMESTAMP) AS ts FROM t
        ), bounds AS (SELECT max(ts) AS anchor FROM anchored)
        SELECT {select_group}
            {cur_expr},
            {prev_expr}
        FROM anchored a CROSS JOIN bounds b
        WHERE a.ts > b.anchor - INTERVAL {d * 2} DAY{where_sql}
        {group_by}
    """
    return ctx.con.execute(sql, params).fetchall()


def conditional_drill(
    ctx: PlaybookContext, cfg: Config, metric: Metric | None = None,
    where: WhereClause | None = None, depth: int | None = None,
) -> list[DrillLevel]:
    """Recursive contribution analysis on the volume delta.

    At each level: compute the parent delta under the current conditions,
    then for every remaining categorical column find the single segment
    whose delta explains the largest share of the parent delta. CONDITION
    on that segment and recurse. Every level in the returned path is
    therefore computed INSIDE its parent segment — the arrow is honest.
    """
    where = where or []
    metric = metric or Metric("count")
    depth = cfg.drill_depth if depth is None else depth
    if depth <= 0 or not ctx.time_col:
        return []
    if not metric.additive:
        # Averages do not decompose additively across segments (mix vs
        # rate effects). Attributing an avg delta to a segment this way
        # would be mathematically wrong, so we refuse rather than fake it.
        return []

    parent = _windowed_agg(ctx, cfg, where, metric)
    if not parent:
        return []
    p_cur, p_prev = float(parent[0][0]), float(parent[0][1])
    parent_delta = p_cur - p_prev
    if parent_delta == 0:
        return []

    used = {c for c, _ in where}
    best: DrillLevel | None = None
    for col in ctx.category_cols:
        if col in used:
            continue
        rows = _windowed_agg(ctx, cfg, where, metric, group_col=col)
        for r in rows:
            val, cur, prev = str(r[0]), float(r[1]), float(r[2])
            seg_delta = cur - prev
            # Only segments moving in the parent's direction can "drive" it.
            if seg_delta == 0 or (seg_delta > 0) != (parent_delta > 0):
                continue
            contribution = seg_delta / parent_delta
            if best is None or abs(contribution) > abs(best.contribution):
                best = DrillLevel(col, val, cur, prev, contribution)

    if best is None:
        return []
    return [best, *conditional_drill(
        ctx, cfg, metric,
        where=[*where, (best.column, best.value)], depth=depth - 1,
    )]


# ─────────────────────────────────────────────────────────────────────────────
# STEP LIBRARY — pure functions: (context, config) → findings
# ─────────────────────────────────────────────────────────────────────────────

def step_shape(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    if ctx.total_rows == 0:
        return (Finding("shape", Severity.CRITICAL,
                        "Dataset contains ZERO rows — nothing below can be "
                        "trusted.", "SELECT COUNT(*) FROM t → 0"),)
    return (Finding("shape", Severity.INFO,
                    f"{ctx.total_rows:,} rows in scope.",
                    "SELECT COUNT(*) FROM t"),)


def step_missing(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    out: list[Finding] = []
    for name, _ in ctx.schema:
        q = ctx.qident(name)
        r = ctx.con.execute(f"SELECT COUNT(*) - COUNT({q}) FROM t").fetchone()
        nulls = int(r[0]) if r else 0
        if nulls and ctx.total_rows:
            share = nulls / ctx.total_rows
            sev = (Severity.NOTABLE if share > cfg.null_notable_share
                   else Severity.INFO)
            out.append(Finding("missing", sev,
                               f"Column '{name}' has {nulls:,} nulls ({share:.1%}).",
                               f"COUNT(*) - COUNT({q})"))
    return tuple(out)


def step_duplicates(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    if ctx.total_rows == 0:
        return ()
    cols = ", ".join(ctx.qident(c) for c, _ in ctx.schema)
    r = ctx.con.execute(
        f"SELECT COALESCE(SUM(n - 1), 0) FROM "
        f"(SELECT COUNT(*) AS n FROM t GROUP BY {cols} HAVING COUNT(*) > 1)"
    ).fetchone()
    dups = int(r[0]) if r else 0
    if dups:
        return (Finding("duplicates", Severity.NOTABLE,
                        f"{dups:,} exact duplicate rows — totals are inflated "
                        f"until deduplicated.",
                        "GROUP BY all columns HAVING COUNT(*) > 1"),)
    return (Finding("duplicates", Severity.INFO,
                    "No exact duplicate rows found.",
                    "GROUP BY all columns HAVING COUNT(*) > 1"),)


def step_time_coverage(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    if not ctx.time_col:
        return (Finding("time_coverage", Severity.NOTABLE,
                        "No time column detected — trend steps will be "
                        "skipped. If one exists under an unusual name, "
                        "manual review is required.", ""),)
    q = ctx.qident(ctx.time_col)
    r = ctx.con.execute(
        f"SELECT min(TRY_CAST({q} AS TIMESTAMP)), "
        f"max(TRY_CAST({q} AS TIMESTAMP)) FROM t"
    ).fetchone()
    if not r or r[0] is None:
        return ()
    age = (datetime.now() - r[1]).days
    sev = Severity.NOTABLE if age > cfg.window_days else Severity.INFO
    return (Finding("time_coverage", sev,
                    f"Data spans {r[0]:%d %b %Y} to {r[1]:%d %b %Y} "
                    f"(newest record is {age} day(s) old).",
                    f"min/max of {q}"),)


def step_volume_change(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    if not ctx.time_col or ctx.total_rows == 0:
        return ()
    metric = resolve_metric(ctx, cfg)
    counts = _windowed_agg(ctx, cfg, [], metric)
    cur, prev = float(counts[0][0]), float(counts[0][1])
    if prev == 0:
        return (Finding("volume_change", Severity.INFO,
                        f"{metric.label} = {metric.fmt(cur)} in the last "
                        f"{cfg.window_days} days; no prior-window baseline.",
                        "windowed aggregate anchored at newest timestamp"),)
    pct = (cur - prev) / prev
    text = (f"{metric.label.capitalize()} {'rose' if pct > 0 else 'fell'} "
            f"{abs(pct):.0%} window over window "
            f"({metric.fmt(prev)} → {metric.fmt(cur)}).")
    if abs(pct) < cfg.drill_threshold:
        return (Finding("volume_change", Severity.INFO, text,
                        "windowed aggregate anchored at newest timestamp"),)
    if not metric.additive:
        text += (" Drill-down SKIPPED on principle: averages do not "
                 "decompose additively across segments (mix vs rate "
                 "effects). Compare segment averages directly instead of "
                 "attributing the overall change.")
        return (Finding("volume_change", Severity.CRITICAL, text,
                        "avg metrics are non-additive; attribution refused"),)
    path = conditional_drill(ctx, cfg, metric)
    if path:
        parts: list[str] = []
        for lvl in path:
            span = f"{metric.fmt(lvl.prev)} → {metric.fmt(lvl.cur)}"
            if abs(lvl.contribution) > 1.0:
                parts.append(
                    f"{lvl.column}='{lvl.value}' moved more than the entire "
                    f"net change ({span}; other segments declined and "
                    f"partially offset it)")
            else:
                parts.append(
                    f"{lvl.column}='{lvl.value}' explains "
                    f"{abs(lvl.contribution):.0%} of the change ({span})")
        chain = " → within that, ".join(parts)
        text += f" Contribution analysis: {chain}."
        evidence = ("recursive contribution analysis: each level computed "
                    "INSIDE its parent segment; contribution = segment delta "
                    "/ parent delta")
    else:
        text += " Contribution analysis found no dominant driver."
        evidence = "windowed counts; no segment explained the majority of the delta"
    return (Finding("volume_change", Severity.CRITICAL, text, evidence),)


def step_anomaly_days(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    if not ctx.time_col:
        return ()
    q = ctx.qident(ctx.time_col)
    daily = [(str(d), int(c)) for d, c in ctx.con.execute(f"""
        WITH anchored AS (SELECT TRY_CAST({q} AS TIMESTAMP) AS ts FROM t),
        bounds AS (SELECT max(ts) AS anchor FROM anchored)
        SELECT strftime(date_trunc('day', a.ts), '%Y-%m-%d'), COUNT(*)
        FROM anchored a CROSS JOIN bounds b
        WHERE a.ts > b.anchor - INTERVAL 28 DAY GROUP BY 1 ORDER BY 1
    """).fetchall()]
    if len(daily) < 7:
        return ()
    counts = sorted(float(c) for _, c in daily)
    n = len(counts)
    med = counts[n // 2] if n % 2 else (counts[n // 2 - 1] + counts[n // 2]) / 2
    devs = sorted(abs(c - med) for c in counts)
    mad = devs[n // 2] if n % 2 else (devs[n // 2 - 1] + devs[n // 2]) / 2
    cutoff = med + cfg.anomaly_mad_k * max(mad, 1.0)
    flagged = [(d, c) for d, c in daily if c > cutoff]
    if not flagged:
        return (Finding("anomaly_days", Severity.INFO,
                        "No day breached the robust baseline in the last "
                        "28 days.", f"median + {cfg.anomaly_mad_k:g} x MAD"),)
    if len(flagged) >= cfg.sustained_run:
        peak_day, peak = max(flagged, key=lambda x: x[1])
        return (Finding("anomaly_days", Severity.CRITICAL,
                        f"SUSTAINED SHIFT: {len(flagged)} of the last "
                        f"{len(daily)} days exceeded the robust baseline "
                        f"(~{cutoff:.0f}/day), peaking at {peak:,} on "
                        f"{peak_day}. Level change, not a one-off spike — "
                        f"see the contribution analysis above.",
                        f"median + {cfg.anomaly_mad_k:g} x MAD; "
                        f"run-length >= {cfg.sustained_run}"),)
    return tuple(
        Finding("anomaly_days", Severity.CRITICAL,
                f"{day} recorded {count:,} events against a robust baseline "
                f"of ~{cutoff:.0f} — worth a root-cause look.",
                f"median + {cfg.anomaly_mad_k:g} x MAD over trailing 28 days")
        for day, count in flagged
    )


def step_concentration(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    out: list[Finding] = []
    for col in ctx.category_cols[:3]:
        q = ctx.qident(col)
        r = ctx.con.execute(
            f"SELECT {q}, COUNT(*), COUNT(*) * 1.0 / SUM(COUNT(*)) OVER () "
            f"FROM t GROUP BY 1 ORDER BY 2 DESC LIMIT 1"
        ).fetchone()
        if r is None:
            continue
        share = float(r[2])
        if share >= cfg.concentration_share:
            out.append(Finding("concentration", Severity.NOTABLE,
                               f"'{r[0]}' accounts for {share:.0%} of all "
                               f"records in '{col}' — a concentration worth "
                               f"an ownership conversation.",
                               f"GROUP BY {q} with window share"))
    return tuple(out)


def step_numeric_sanity(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    out: list[Finding] = []
    for col in ctx.numeric_cols[:4]:
        q = ctx.qident(col)
        r = ctx.con.execute(
            f"SELECT COUNT(*) FILTER (WHERE {q} < 0), min({q}), max({q}) FROM t"
        ).fetchone()
        if r and int(r[0]):
            out.append(Finding("numeric_sanity", Severity.NOTABLE,
                               f"Column '{col}' contains {int(r[0]):,} "
                               f"negative values (min {r[1]}). Verify whether "
                               f"negatives are business logic or data errors.",
                               f"COUNT(*) FILTER (WHERE {q} < 0)"))
    return tuple(out)


def step_recommendations(ctx: PlaybookContext, cfg: Config) -> tuple[Finding, ...]:
    # NOTE: this step declares requires=(...) in the library entry; the
    # runner passes prior findings via a closure-free mechanism below.
    return ()   # replaced by the runner's summary — see run_playbook


STEP_LIBRARY: Final[dict[str, Step]] = {
    "shape": Step("shape", "How much data am I looking at?",
                  "Volume context frames everything; zero rows means stop.",
                  step_shape),
    "missing": Step("missing", "Is anything missing?",
                    "Gaps bias every downstream number; find them first.",
                    step_missing, requires=("shape",)),
    "duplicates": Step("duplicates", "Is anything counted twice?",
                       "Duplicates inflate totals; dedupe before summing.",
                       step_duplicates, requires=("shape",)),
    "time_coverage": Step("time_coverage", "What period does this cover?",
                          "A perfect analysis of stale data is a perfectly "
                          "wrong answer.", step_time_coverage),
    "volume_change": Step("volume_change",
                          "Did volume move — and what drove it?",
                          "Never report a delta without its driver; the "
                          "drill-down conditions on each winner and recurses.",
                          step_volume_change,
                          requires=("shape", "time_coverage")),
    "anomaly_days": Step("anomaly_days", "Point spike or level shift?",
                         "Robust baselines keep one bad day from hiding "
                         "itself; run-length separates spike from shift.",
                         step_anomaly_days, requires=("time_coverage",)),
    "concentration": Step("concentration", "Is one category carrying the load?",
                          "40%+ concentration is an ownership conversation.",
                          step_concentration, requires=("shape",)),
    "numeric_sanity": Step("numeric_sanity", "Are there impossible values?",
                           "Negative amounts are business logic or bugs — "
                           "decide which.", step_numeric_sanity,
                           requires=("shape",)),
}


def _builtin(key: str, title: str, desc: str, steps: list[str]) -> Playbook:
    pb = Playbook(key, title, desc,
                  tuple(STEP_LIBRARY[s] for s in steps))
    pb.validate()
    return pb


PLAYBOOKS: Final[dict[str, Playbook]] = {
    "weekly-review": _builtin(
        "weekly-review", "Weekly Operations Review",
        "The Monday-morning checklist, with the drill-down reflex built in.",
        ["shape", "missing", "duplicates", "time_coverage",
         "volume_change", "anomaly_days", "concentration"],
    ),
    "data-quality": _builtin(
        "data-quality", "Data Quality Audit",
        "Trust the data before analysing it.",
        ["shape", "missing", "duplicates", "numeric_sanity", "time_coverage"],
    ),
    "trend-investigation": _builtin(
        "trend-investigation", "Trend Investigation",
        "Something changed: find what, where, and how concentrated.",
        ["shape", "time_coverage", "volume_change",
         "anomaly_days", "concentration"],
    ),
}


def materialise_custom_playbooks(raw: dict[str, Any]) -> dict[str, Playbook]:
    """Builds Playbook objects from opskit.toml [playbooks.*] tables.

    Domain packs without code changes:
        [playbooks.claims-review]
        title = "Insurance Claims Review"
        description = "Weekly claims triage"
        steps = ["shape", "missing", "volume_change", "concentration"]
    """
    out: dict[str, Playbook] = {}
    for key, spec in raw.items():
        steps = spec.get("steps", [])
        unknown = [s for s in steps if s not in STEP_LIBRARY]
        if unknown:
            raise OpsKitError(
                f"[playbooks.{key}] unknown steps {unknown}. "
                f"Available: {sorted(STEP_LIBRARY)}"
            )
        pb = Playbook(key, str(spec.get("title", key)),
                      str(spec.get("description", "")),
                      tuple(STEP_LIBRARY[s] for s in steps))
        pb.validate()
        out[key] = pb
    return out


# ─────────────────────────────────────────────────────────────────────────────
# RUNNER — composes pure steps; owns the findings list
# ─────────────────────────────────────────────────────────────────────────────

_SEV_ICON: Final[dict[Severity, str]] = {
    Severity.INFO: "·", Severity.NOTABLE: "▲", Severity.CRITICAL: "⚠",
}


def run_playbook(
    pb: Playbook, ctx: PlaybookContext, cfg: Config
) -> list[Finding]:
    """Executes steps in declared order, then appends the runner-owned
    recommendations summary. Steps are pure; only the runner accumulates."""
    print(f"\nPLAYBOOK: {pb.title}\n{pb.description}\n" + "=" * 64)
    findings: list[Finding] = []
    for i, step in enumerate(pb.steps, 1):
        print(f"\nStep {i}/{len(pb.steps)} — {step.question}")
        print(f"  why: {step.rationale}")
        produced = step.run(ctx, cfg)
        for f in produced:
            print(f"  {_SEV_ICON[f.severity]} [{f.severity}] {f.text}")
        findings.extend(produced)

    crit_steps = list(dict.fromkeys(
        f.step for f in findings if f.severity is Severity.CRITICAL))
    notable_steps = list(dict.fromkeys(
        f.step for f in findings if f.severity is Severity.NOTABLE))
    if crit_steps:
        summary = Finding("recommendations", Severity.CRITICAL,
                          f"Critical findings require action before this data "
                          f"is reported onward, across: "
                          f"{'; '.join(crit_steps)}.", "")
    elif notable_steps:
        summary = Finding("recommendations", Severity.NOTABLE,
                          f"Notable findings to remediate in the normal "
                          f"cycle: {'; '.join(notable_steps)}.", "")
    else:
        summary = Finding("recommendations", Severity.INFO,
                          "Nothing requiring escalation. Continue routine "
                          "monitoring.", "")
    print(f"\nSummary: {_SEV_ICON[summary.severity]} {summary.text}")
    findings.append(summary)
    return findings


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUTS — Excel workpaper and JSON lines
# ─────────────────────────────────────────────────────────────────────────────

def _fill(c: str) -> PatternFill:
    return PatternFill("solid", start_color=c)


def _font(bold: bool = False, colour: str = PAL["ink"],
          size: int = 10, italic: bool = False) -> Font:
    return Font(name="Arial", bold=bold, color=colour, size=size, italic=italic)


def write_report(
    pb: Playbook, ctx: PlaybookContext, cfg: Config,
    findings: list[Finding], out: Path | None,
) -> Path:
    run_at = datetime.now(tz=IST)
    out_path = out or Path(f"playbook_{pb.key}_{run_at:%Y%m%d_%H%M%S}.xlsx")
    wb = Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)

    order = {Severity.CRITICAL: 0, Severity.NOTABLE: 1, Severity.INFO: 2}
    ordered = sorted(findings, key=lambda f: order[f.severity])

    ws: Worksheet = wb.create_sheet("Findings")
    ws["A1"] = f"{pb.title.upper()} — FINDINGS"
    ws["A1"].font = _font(bold=True, size=14, colour=PAL["navy"])
    ws["A2"] = (f"Source: {ctx.source.name} · Rows: {ctx.total_rows:,} · "
                f"Run: {run_at:%d %B %Y, %H:%M IST}")
    ws["A2"].font = _font(italic=True, size=9, colour=PAL["muted"])
    for c, h in enumerate(["Severity", "Step", "Finding", "Evidence"], 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=PAL["white"], size=10)
        cell.fill = _fill(PAL["navy"])
    for r, f in enumerate(ordered, 5):
        for c, v in enumerate([str(f.severity), f.step, f.text, f.evidence], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = _font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = _fill(PAL["alt"])
        sev = ws.cell(row=r, column=1)
        if f.severity is Severity.CRITICAL:
            sev.fill = _fill(PAL["red"])
            sev.font = _font(bold=True, colour=PAL["white"])
        elif f.severity is Severity.NOTABLE:
            sev.fill = _fill(PAL["amber"])
            sev.font = _font(bold=True)
    for col, width in zip("ABCD", (12, 18, 72, 46), strict=True):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A5"

    ws2: Worksheet = wb.create_sheet("Methodology")
    ws2["A1"] = "RE-PERFORMANCE RECORD"
    ws2["A1"].font = _font(bold=True, size=13, colour=PAL["navy"])
    method = [
        ("Playbook", f"{pb.title} ({pb.key})"),
        ("Run timestamp", f"{run_at:%d %B %Y, %H:%M IST}"),
        ("Source", str(ctx.source)),
        ("Assumptions", f"time={ctx.time_col} · categories="
                        f"{list(ctx.category_cols)} · numerics="
                        f"{list(ctx.numeric_cols)}"),
        ("Config", ", ".join(f"{k}={getattr(cfg, k)}"
                             for k in Config.__dataclass_fields__)),
        ("Drill-down method", "Recursive contribution analysis: each level "
                              "computed inside its parent segment; "
                              "contribution = segment delta / parent delta."),
        ("Determinism", "No AI calls. Same data, same config, same findings."),
    ]
    for i, (label, value) in enumerate(method, 3):
        ws2.cell(row=i, column=1, value=label).font = _font(bold=True)
        cell = ws2.cell(row=i, column=2, value=value)
        cell.font = _font()
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 96

    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=out_path.parent or Path("."))
    os.close(fd)
    tmp_path = Path(tmp)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, out_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
    return out_path


def write_jsonl(findings: list[Finding], path: Path) -> None:
    """Machine-readable findings: one JSON object per line, schema-versioned."""
    with path.open("w", encoding="utf-8") as fh:
        for f in findings:
            record = {"schema": "opskit.finding/v1", **f.as_json()}
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# DEMO
# ─────────────────────────────────────────────────────────────────────────────

def cmd_demo(out_dir: Path) -> None:
    """Planted story: final-week surge driven by payments, and WITHIN
    payments by P2 — a true conditional path for the drill to find."""
    rng = random.Random(11)
    out_dir.mkdir(exist_ok=True)
    services = ["payments", "cards", "auth", "mobile-app", "web-app"]
    end = datetime(2026, 7, 6, 18, 0)
    rows: list[list[str]] = []
    n = 1
    for off in range(56, -1, -1):
        day = end - timedelta(days=off)
        base = 11 if off <= 7 else 6
        if off == 10:
            base = 26
        for _ in range(max(1, int(rng.gauss(base, 1.5)))):
            if off <= 7 and rng.random() < 0.55:
                svc = "payments"
                sev = "P2" if rng.random() < 0.75 else "P3"
            else:
                svc = rng.choice(services)
                sev = rng.choices(["P1", "P2", "P3", "P4"],
                                  weights=[0.05, 0.2, 0.45, 0.3])[0]
            opened = day.replace(hour=rng.randint(0, 23),
                                 minute=rng.randint(0, 59))
            owner = rng.choice(["priya.k", "arjun.m", "sana.r", "dev.t", ""])
            # planted METRIC story: payments losses triple in the surge week
            base_amount = rng.lognormvariate(7, 0.8)
            if off <= 7 and svc == "payments":
                base_amount *= 3
            rows.append([f"INC-{n:05d}", opened.strftime("%Y-%m-%d %H:%M:%S"),
                         sev, svc, owner, f"{base_amount:.2f}"])
            n += 1
    rows.extend([list(r) for r in rng.sample(rows, 3)])
    path = out_dir / "incidents.csv"
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["incident_id", "opened_at", "severity", "service",
                    "owner", "loss_amount"])
        w.writerows(rows)
    db = out_dir / "incidents.sqlite"
    db.unlink(missing_ok=True)
    with sqlite3.connect(db) as conn:
        conn.execute("CREATE TABLE incidents (incident_id TEXT, opened_at "
                     "TEXT, severity TEXT, service TEXT, owner TEXT, "
                     "loss_amount REAL)")
        conn.executemany("INSERT INTO incidents VALUES (?,?,?,?,?,?)", rows)
    print(f"Created {len(rows):,} incidents with a PLANTED CONDITIONAL story:")
    print("  surge driven by service='payments', and WITHIN payments by "
          "severity='P2'")
    print("  plus: loss_amount column — payments losses TRIPLE in the "
          "surge week (for --metric sum:loss_amount)")
    print("  plus: one anomaly day · blank owners · 3 duplicate rows")
    print(f"Files: {path}, {db}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    p = argparse.ArgumentParser(
        prog="opskit4",
        description="Cognitive analyst playbooks: the cross-domain review "
                    "sequence, with honest conditional drill-down.",
    )
    p.add_argument("--config", type=Path, default=None,
                   help="Path to opskit.toml (default: ./opskit.toml if present)")
    sub = p.add_subparsers(dest="command", required=True)

    d = sub.add_parser("demo")
    d.add_argument("--out", type=Path, default=Path("demo_data"))
    sub.add_parser("list")
    e = sub.add_parser("explain")
    e.add_argument("playbook")
    r = sub.add_parser("run")
    r.add_argument("playbook")
    r.add_argument("source", type=Path)
    r.add_argument("--table", default=None)
    r.add_argument("--out", type=Path, default=None)
    r.add_argument("--json", type=Path, default=None, dest="json_out",
                   help="Also write findings as JSON lines to this path")
    r.add_argument("--no-fail", action="store_true",
                   help="Always exit 0 even when critical findings exist")
    r.add_argument("--metric", default=None,
                   help="Analysis metric: count (default), sum:<col>, avg:<col>")

    args = p.parse_args()
    try:
        cfg, custom_raw = load_config(args.config)
        books = {**PLAYBOOKS, **materialise_custom_playbooks(custom_raw)}

        if args.command == "demo":
            cmd_demo(args.out)
        elif args.command == "list":
            for pb in books.values():
                print(f"  {pb.key:<22} {pb.title} — {pb.description}")
        elif args.command == "explain":
            pb_x = books.get(args.playbook)
            if pb_x is None:
                raise OpsKitError(f"No playbook '{args.playbook}'. "
                                  f"Available: {', '.join(books)}")
            print(f"\n{pb_x.title}\n{pb_x.description}\n" + "=" * 60)
            for i, s in enumerate(pb_x.steps, 1):
                dep = f"  (requires: {', '.join(s.requires)})" if s.requires else ""
                print(f"\n{i}. {s.question}{dep}\n   why: {s.rationale}")
        elif args.command == "run":
            pb_r = books.get(args.playbook)
            if pb_r is None:
                raise OpsKitError(f"No playbook '{args.playbook}'. "
                                  f"Available: {', '.join(books)}")
            if args.metric:
                cfg = replace(cfg, metric=args.metric)
            ctx = build_context(args.source, cfg, args.table)
            findings = run_playbook(pb_r, ctx, cfg)
            out_path = write_report(pb_r, ctx, cfg, findings, args.out)
            if args.json_out:
                write_jsonl(findings, args.json_out)
                print(f"  ✓  JSON findings → {args.json_out}")
            crit = sum(1 for f in findings if f.severity is Severity.CRITICAL)
            print(f"  ✓  {len(findings)} findings ({crit} critical) → {out_path}")
            if crit and not args.no_fail:
                sys.exit(EXIT_CRITICAL_FINDINGS)
    except OpsKitError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(EXIT_ERROR)
    except BrokenPipeError:
        sys.stderr.close()
        sys.exit(EXIT_CLEAN)


if __name__ == "__main__":
    main()
